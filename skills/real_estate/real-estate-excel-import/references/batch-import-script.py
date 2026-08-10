# 批量导入脚本模板
# v1.1.0 - 修复字段映射错误：
#   - 景观不再误填 district（之前导致 423 条房源区域全部错误）
#   - 景观改存 tags，保留在 title
#   - renovation 不再硬编码"精装"（避免编造数据）
#   - 出租房源价格单位修正：Excel 是"元/月"，系统 price 是"万元"，需 ÷10000

## execute_code 中的数据处理模板

```python
from openpyxl import load_workbook
import json, re

def parse_layout(s):
    if not s: return 1, 0
    s = str(s).strip()
    if '开间' in s: return 1, 0
    if '一房一厅' in s: return 1, 1
    if '两房一厅' in s: return 2, 1
    if '三房两厅' in s: return 3, 2
    if '四房两厅' in s: return 4, 2
    nums = re.findall(r'\d+', s)
    return (int(nums[0]), int(nums[1])) if len(nums)>=2 else (int(nums[0]) if nums else 1, 0)

def safe_float(val):
    if val is None: return 0
    s = str(val).strip()
    if '万' in s: return float(s.replace('万',''))*10000
    if s in ('','0','/'): return 0
    try: return float(s)
    except: return 0

wb = load_workbook(file_path, data_only=True)
all_props = []

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    is_rental = '出租' in sheet_name

    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, values_only=True):
        if all(c is None for c in row) or row[2] is None:
            continue

        room_num = str(row[5]) if row[5] else ""
        layout = str(row[6]) if row[6] else ""
        area = safe_float(row[7])
        landscape = str(row[16]) if len(row)>16 and row[16] else ""
        status = str(row[20]) if len(row)>20 and row[20] else ""
        notes = str(row[21]) if len(row)>21 and row[21] else ""

        if is_rental:
            # Excel 出租价是"元/月"（如 1000 = 1000元/月），系统 price 单位是"万元"
            # 6个月租价在 col10，年租价在 col12；取非空值，单位统一转为万元
            price_raw = safe_float(row[12]) or safe_float(row[10])
            prop_type = 'rental'
            price_wan = round(price_raw / 10000, 4) if price_raw > 0 else 0
        else:
            # Excel 出售总价是"元"（如 400000 = 40万），转万元
            price_raw = safe_float(row[12])
            prop_type = 'second_hand'
            price_wan = round(price_raw / 10000, 2) if price_raw > 10000 else round(price_raw, 2)

        rooms, halls = parse_layout(layout)

        title = f"{room_num} {layout} {area}㎡"
        if landscape: title += f" {landscape}"

        # 景观是房源卖点，存入 tags（不占 district/区域字段）
        tags = f"景观:{landscape}" if landscape else ""

        if (prop_type == 'second_hand' and price_wan > 0 and area > 0) or \
           (prop_type == 'rental' and area > 0):
            all_props.append({
                'title': title, 'price': price_wan, 'area': area,
                'rooms': rooms, 'halls': halls, 'property_type': prop_type,
                # district/community/renovation 无真实数据时留空，禁止编造
                'district': '', 'community': '', 'renovation': '',
                'tags': tags, 'notes': notes, 'source_sheet': sheet_name
            })

# 保存JSON
with open('/tmp/properties_to_import.json', 'w', encoding='utf-8') as f:
    json.dump(all_props, f, ensure_ascii=False, indent=2)

print(f"共 {len(all_props)} 条待导入")
```

## delegate_task 子代理目标

```
从/tmp/properties_to_import.json批量导入房源。
跳过前N条（已导入），对每条调用add_property。
每10条打印进度，完成后汇报成功/失败数和property_stats统计。
```

## 数据修复（历史脏数据）

如果此前用旧版脚本导入过数据（district 被填成"园区/海景"、出租价格未转万元），
用以下 SQL 在服务器上修复（见 git 提交说明）：
- 景观从 district 移到 tags：`UPDATE re_properties SET tags = '景观:'||district WHERE district IN ('园区','海景','园林');`
- 区域置空：`UPDATE re_properties SET district = '' WHERE district IN ('园区','海景','园林');`
- 出租价格转万元：`UPDATE re_properties SET price = price/10000, unit_price = unit_price/10000 WHERE property_type='rental' AND price > 100;`
